#Version 5/30/25
import os, sys
import pandas as pd
import numpy as np
from openpyxl import load_workbook
import itertools

path_libs = os.getcwd() + os.sep + 'libs' + os.sep
if not path_libs in sys.path: sys.path.append(path_libs)
import parsetables

"""
================================================================================
ProjectTables Class -- this can be initialized as tbls to manage all data
tables for a project. __init__ instances a Table class for each table and
initializes lists column name mapping and for preflight checks

JDL 9/3/24
================================================================================
"""
class ProjectTables():
    """
    Collection of imported or generated data tables for a project
    JDL 9/26/24; Modified 5/28/25 add UseTblInfo flag
    """
    def __init__(self, files, UseTblInfo=False, UseColInfo=False, IsPrint=False):
        """
        Instance attributes including Table instances
        """
        self.files = files
        self.UseTblInfo = UseTblInfo
        self.UseColInfo = UseColInfo
        self.IsPrint = IsPrint

        # instance self.tbl_info and import from files.pf_col_info
        if self.UseTblInfo:
            #self.ImportTblInfoDf()
            pass

        if self.UseColInfo:
            self.col_info = ColumnInfo(files, IsInit=True, IsPrint=self.IsPrint)

        #Instance project-specific tables if any
        #self.InstanceTblObjs()
        
    def InstanceTblObjs(self):
        """
        Instance project-specific tables 
        JDL 4/9/25
        """
        # Add project-specific table instantiations here
        # Typical is do define dImportParams and dParseParams dicts and pass as
        # arguments to Table(name, dImportParams, dParseParams, tbls.col_info) along
        # with optional col_info instance
        pass

class Table():
    """
    Attributes for a data table including import instructions and other
    metadata. Table instances are attributes of ProjectTables Class
    JDL Modified 4/8/25 refactor to fully use dImportParams and dParseParams
        5/28/25 to add col_info attribute
    """
    def __init__(self, name, dImportParams=None, dParseParams=None, col_info=None):
                
        self.name = name #Table name

        # Dicts of import and parsing parameters
        self.dImportParams = dImportParams or {}
        self.dParseParams = dParseParams or {'parse_type':'none'}
        self.df_raw = pd.DataFrame() # temp raw data in .lst_dfs iteration
        self.df = pd.DataFrame()

        # optionally create column info df (subset of all col_info) for this table
        self.dfColInfo = None
        if not col_info is None: self.SetTblColInfo(col_info)

        # Temp variables for looping through files
        self.pf = None
        self.sht = None
        self.lst_dfs = None
        self.sht_type = None
        self.is_unstructured = None
        self.lst_dfs = None

    def SetTblColInfo(self, col_info):
        """
        Subset overall column info for this table
        JDL 5/28/25
        """
        self.dfColInfo = col_info.df[col_info.df['tbl_name'] == self.name].copy()
        """
    ================================================================================
    ParseRawData Procedure
    * If imported data are not rows/cols structured, parse .lst_dfs to .df using
    dParseParams inputs to guide parsing
    * all parsing classes have ParseDfRawProcedure that populates parse.df
    JDL 4/21/25; updated 5/30/25
    ================================================================================
    """
    def ParseRawData(self):
        """
        Procedure to parse raw data for a given Table instance
        Updated 5/30/25
        """
        for self.df_raw in self.lst_dfs:

            # instance parse class with tbl (aka self) as argument and parse
            parse = getattr(parsetables, self.dParseParams['parse_type'])(self)
            parse.ParseDfRawProcedure()
                
            # Concatenate parsed data to tbl.df
            self.df = pd.concat([self.df, parse.df], ignore_index=True)

    """
    ================================================================================
    ImportToTblDf Procedure
    JDL 4/10/25 Rewritten to allow multisheet Excel and separate ingest/parse
    ================================================================================
    """
    def ImportToTblDf(self, lst_files=None):
        """
        Procedure to import file(s) + sheet(s) to self.df (structured rows/cols)
        or self.lst_dfs (unstructured) using .dImportParams and .dParseParams to
        set options

        Can directly specify lst_files as arg or as dImportParams['lst_files']
        Refactored JDL 4/10/25; Add IsAddFilenameCol option 4/29/25
        """

        # Set lst_files based on dImportParams['lst_files'] or input arg    
        lst_files = self.SetLstFiles(lst_files)

        # Set file ingest parameters based on dImportParams or on default values
        self.SetFileIngestParams()

        # initialize list df's and temp df (temp if structured; or for parsing later)
        self.lst_dfs = []
        self.df_temp = pd.DataFrame()

        # Loop over input list of files to ingest
        for self.pf in lst_files:

            # Read from Excel single/multiple sheets self.pf; append to lst_dfs
            if self.dImportParams['ftype'] == 'excel':
                self.SetLstSheets()
                self.ReadExcelFileSheets()

            # Read from CSV self.pf; append to lst_dfs
            elif self.dImportParams['ftype'] == 'csv':
                self.ReadCSVFile()

            # Read from feather self.pf; append to lst_dfs
            elif self.dImportParams['ftype'] == 'feather':
                pass

        #Concat if rows/cols aka structured (e.g. no parsing needed) and list is non-empty
        if not self.is_unstructured and self.lst_dfs:
            self.df = pd.concat(self.lst_dfs, ignore_index=True)
            self.lst_dfs = []

    def SetLstFiles(self, lst_files):
        """
        Set lst_files based on input and dImportParams.
        """
        # If lst_files is not specified, use dImportParams['lst_files']
        if lst_files is None: lst_files = self.dImportParams['lst_files']

        # Convert to list if a single file is provided
        if not isinstance(lst_files, list): lst_files = [lst_files]

        # Optionally prepend import_path to each file name
        if 'import_path' in self.dImportParams:
            lst_files = [self.dImportParams['import_path'] + f for f in lst_files]
        return lst_files

    def SetFileIngestParams(self):
        """
        Set Table attributes for the current file 
        (concise vs referencing dict items and also factors in default vals if
        dict item not specified)
        JDL 4/10/25
        """
        self.is_unstructured = self.SetParseParam(False, 'is_unstructured')
        self.n_skip_rows = self.SetParseParam(0, 'n_skip_rows')
        self.parse_type = self.SetParseParam('none', 'parse_type')
        if self.dImportParams['ftype'] == 'excel':
            self.sht_type = self.SetImportParam('single', 'sht_type')

    def SetImportParam(self, valDefault, param_name):
        """
        Set default or non-default import parameter
        JDL 4/9/25
        """
        val = valDefault
        if param_name in self.dImportParams: val = self.dImportParams[param_name]
        return val

    def SetParseParam(self, valDefault, param_name):
        """
        Set default or non-default parsing parameter
        JDL 4/9/25
        """
        val = valDefault
        if param_name in self.dParseParams: val = self.dParseParams[param_name]
        return val

    def SetLstSheets(self):
        """
        Set .lst_sheets based on sht_type and sht in dImportParams
        (Called within iteration with self.pf file)
        JDL 4/10/25; Updated 5/30/25
        """
        self.lst_sheets = []

        if self.sht_type == 'single':

            # Set sheet name to either specified or 0 (e.g. first sheet)
            self.lst_sheets = [self.dImportParams.get('sht', 0)]

            # If sheet name 0, reset to first sheet name (must open to see sht names)
            if self.lst_sheets[0] == 0:
                wb = load_workbook(filename=self.pf, read_only=True)
                self.lst_sheets[0] = wb.sheetnames[0]
                wb.close()

        elif self.sht_type == 'all':
            wb = load_workbook(filename=self.pf, read_only=True)
            self.lst_sheets = wb.sheetnames
            wb.close()
        
        elif self.sht_type == 'list':
            pass

        elif self.sht_type == 'regex':
            pass

        elif self.sht_type == 'startswith':
            pass

        elif self.sht_type == 'endswith':
            pass

        elif self.sht_type == 'contains':
            pass

    def ReadExcelFileSheets(self):
        """
        Loop through sheets in lst_sheets and read their data
        JDL 4/10/25
        """
        for self.sht in self.lst_sheets:
            self.ReadExcelSht()
            self.lst_dfs.append(self.df_temp)
            self.df_temp = pd.DataFrame()

    def ReadExcelSht(self):
        """
        Read data from the current sheet into a temporary DataFrame.
        """
        if self.is_unstructured:
            #self.df_temp = self.ImportExcelRaw()
            self.df_temp = pd.read_excel(self.pf, sheet_name=self.sht, header=None)

            # Negate Pandas inferring float data type for integers and NaNs for blanks
            if 'import_dtype' in self.dParseParams and self.dParseParams['import_dtype'] == str:
                self.df_temp = self.df_temp.astype(object)
                self.df_temp = self.df_temp.map(lambda x: None if pd.isna(x) \
                    else str(int(x)) if isinstance(x, float) and x.is_integer() else str(x))
        else:
            self.df_temp = pd.read_excel(self.pf, sheet_name=self.sht, skiprows=self.n_skip_rows)

    def ReadCSVFile(self):
        """
        Import current CSV file into a temporary df and append to lst_dfs
        JDL 4/10/25
        """
        if self.is_unstructured:
            # Read CSV without treating first row as headers
            self.df_temp = pd.read_csv(self.pf, header=None)
        else:
            # Read CSV with optional skiprows
            self.df_temp = pd.read_csv(self.pf, skiprows=self.n_skip_rows)

        # Append temp df to lst_dfs and re-initialize
        self.lst_dfs.append(self.df_temp)
        self.df_temp = pd.DataFrame()

class CheckInputs:
    """
    Check the tbls dataframes for errors
    (dummy initialization of preflight check)
    """
    def __init__(self, tbls, IsPrint=True):
        self.tbls = tbls
        self.IsPrint = IsPrint

        #preflight.CheckDataFrame Class --instanced as needed in methods below
        self.ckdf = None    

